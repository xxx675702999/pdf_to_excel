import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import numpy as np
import cv2
from pdf2image import convert_from_path
import paddleocr
import os
import tempfile
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import unicodedata

class ImageProcessor:
    @staticmethod
    def process_files(file_paths, output_size=(1600, 1200)):
        """处理输入文件并返回图片列表及元数据"""
        images = []
        file_info = []
        for path in file_paths:
            if path.lower().endswith('.pdf'):
                with tempfile.TemporaryDirectory() as temp_dir:
                    poppler_path = r"poppler/Library/bin"  # 根据实际路径修改
                    pages = convert_from_path(
                        path,
                        dpi=300,
                        output_folder=temp_dir,
                        poppler_path=poppler_path
                    )
                    for i, page in enumerate(pages):
                        img = page.resize(output_size, Image.Resampling.LANCZOS)
                        images.append(img)
                        file_info.append({
                            'filename': os.path.basename(path),
                            'filepath': path,
                            'page': i + 1
                        })
            else:
                img = Image.open(path)
                img = img.resize(output_size, Image.Resampling.LANCZOS)
                images.append(img)
                file_info.append({
                    'filename': os.path.basename(path),
                    'filepath': path,
                    'page': None
                })
        return images, file_info


class InvoiceProcessorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("小苏专用发票识别系统❤")

        # 初始化变量
        self.images = []
        self.file_info = []
        self.current_image_index = 0
        self.regions = []  # 存储原始坐标区域
        self.rect_ids = []  # 存储画布矩形对象ID
        self.scale_factor = (1, 1)
        self.output_size = (1600, 1200)

        # 初始化OCR
        self.ocr = paddleocr.PaddleOCR(use_gpu=False,cls_model_dir=".\ocrModel\cls\ch",det_model_dir=".\ocrModel\det\ch",rec_model_dir=".\ocrModel\\rec\ch",use_angle_cls=True, lang='ch')

        # 创建界面
        self.create_widgets()

    def create_widgets(self):
        # 在工具栏添加新按钮
        toolbar = tk.Frame(self.root)
        toolbar.pack(fill=tk.X)

        # 保持原有按钮...
        self.btn_process_all = ttk.Button(toolbar, text="批量识别全部", command=self.process_all)
        self.btn_process_all.pack(side=tk.LEFT, padx=5)

        self.btn_process_single = ttk.Button(toolbar, text="识别当前文件", command=self.process_current)
        self.btn_process_single.pack(side=tk.LEFT, padx=5)
        # 主布局框架
        main_frame = tk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 左侧文件列表区
        left_panel = tk.Frame(main_frame)
        main_frame.add(left_panel, minsize=200)

        # 文件列表
        tk.Label(left_panel, text="文件列表", font=('微软雅黑', 10, 'bold')).pack(pady=5)
        self.file_listbox = tk.Listbox(
            left_panel,
            selectmode=tk.SINGLE,
            height=20,
            font=('微软雅黑', 9)
        )
        self.file_listbox.pack(fill=tk.BOTH, expand=True, padx=5)
        self.file_listbox.bind('<<ListboxSelect>>', self.on_file_select)

        # 右侧主操作区
        right_panel = tk.Frame(main_frame)
        main_frame.add(right_panel, minsize=600)

        # 工具栏
        toolbar = tk.Frame(right_panel)
        toolbar.pack(fill=tk.X)

        ttk.Button(toolbar, text="打开文件", command=self.open_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="撤销区域", command=self.undo_region).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="上一张", command=lambda: self.change_image(-1)).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="下一张", command=lambda: self.change_image(1)).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="开始识别", command=self.process_all).pack(side=tk.LEFT, padx=5)

        # 图像显示区
        self.canvas_frame = tk.Frame(right_panel)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(self.canvas_frame, bg='gray')
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 滚动条
        self.scroll_y = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=self.scroll_y.set)

        # 区域列表
        region_frame = tk.Frame(right_panel)
        region_frame.pack(fill=tk.Y, side=tk.RIGHT, padx=5)

        tk.Label(region_frame, text="区域列表", font=('微软雅黑', 10, 'bold')).pack()
        self.region_list = tk.Listbox(
            region_frame,
            width=35,
            height=20,
            font=('微软雅黑', 9)
        )
        self.region_list.pack(fill=tk.BOTH, expand=True)

        # 绑定事件
        self.canvas.bind("<ButtonPress-1>", self.start_rectangle)
        self.canvas.bind("<B1-Motion>", self.draw_rectangle)
        self.canvas.bind("<ButtonRelease-1>", self.save_rectangle)

    def open_files(self):
        file_types = [('PDF/图像文件', '*.pdf *.jpg *.jpeg *.png')]
        file_paths = filedialog.askopenfilenames(filetypes=file_types)
        if not file_paths:
            return

        try:
            self.images, self.file_info = ImageProcessor.process_files(file_paths, self.output_size)
            self.current_image_index = 0
            self.update_file_list()
            self.show_image()
        except Exception as e:
            messagebox.showerror("错误", f"文件处理失败: {str(e)}")

    def update_file_list(self):
        """更新文件列表显示"""
        self.file_listbox.delete(0, tk.END)
        for info in self.file_info:
            if info['page']:
                display_name = f"{info['filename']} (第{info['page']}页)"
            else:
                display_name = info['filename']
            self.file_listbox.insert(tk.END, display_name)

    def on_file_select(self, event):
        """文件列表选择事件处理"""
        selection = self.file_listbox.curselection()
        if selection:
            self.current_image_index = selection[0]
            self.show_image()

    def show_image(self):
        """显示当前图片并重绘所有区域"""
        if not self.images:
            return

        # 更新文件列表选择状态
        self.file_listbox.selection_clear(0, tk.END)
        self.file_listbox.selection_set(self.current_image_index)
        self.file_listbox.see(self.current_image_index)

        # 清除旧绘图
        self.canvas.delete("all")
        self.rect_ids.clear()

        # 获取当前图片
        img_pil = self.images[self.current_image_index]

        # 计算缩放比例
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        img_width, img_height = img_pil.size
        ratio = min(canvas_width / img_width,
                    canvas_height / img_height) if canvas_width > 0 and canvas_height > 0 else 1
        new_size = (int(img_width * ratio), int(img_height * ratio))

        # 保存缩放比例
        self.scale_factor = (img_width / new_size[0], img_height / new_size[1])

        # 显示图片
        img_resized = img_pil.resize(new_size, Image.Resampling.LANCZOS)
        self.tk_image = ImageTk.PhotoImage(img_resized)
        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.tk_image)
        self.canvas.config(scrollregion=(0, 0, new_size[0], new_size[1]))

        # 重绘所有区域
        for region in self.regions:
            x0, y0, x1, y1 = region
            # 转换到画布坐标
            x0_canvas = x0 / self.scale_factor[0]
            y0_canvas = y0 / self.scale_factor[1]
            x1_canvas = x1 / self.scale_factor[0]
            y1_canvas = y1 / self.scale_factor[1]
            rect_id = self.canvas.create_rectangle(
                x0_canvas, y0_canvas, x1_canvas, y1_canvas,
                outline='red', width=2
            )
            self.rect_ids.append(rect_id)

    def undo_region(self):
        """撤销最后一个区域"""
        if self.regions:
            self.regions.pop()
            self.region_list.delete(tk.END)
            self.show_image()  # 重绘更新

    def change_image(self, delta):
        """切换图片"""
        if not self.images:
            return

        new_index = self.current_image_index + delta
        if 0 <= new_index < len(self.images):
            self.current_image_index = new_index
            self.show_image()

    def start_rectangle(self, event):
        self.rect_start = (event.x, event.y)
        self.rect_id = None

    def draw_rectangle(self, event):
        if self.rect_id:
            self.canvas.delete(self.rect_id)

        x0, y0 = self.rect_start
        x1, y1 = event.x, event.y
        self.rect_id = self.canvas.create_rectangle(x0, y0, x1, y1, outline='red', width=2)

    def save_rectangle(self, event):
        x0, y0 = self.rect_start
        x1, y1 = event.x, event.y

        # 转换为原始图像坐标
        x0_raw = int(x0 * self.scale_factor[0])
        y0_raw = int(y0 * self.scale_factor[1])
        x1_raw = int(x1 * self.scale_factor[0])
        y1_raw = int(y1 * self.scale_factor[1])

        # 确保坐标顺序正确
        x0_raw, x1_raw = sorted([x0_raw, x1_raw])
        y0_raw, y1_raw = sorted([y0_raw, y1_raw])

        self.regions.append((x0_raw, y0_raw, x1_raw, y1_raw))
        self.region_list.insert(tk.END, f"区域{len(self.regions)}: ({x0_raw}, {y0_raw}) - ({x1_raw}, {y1_raw})")

    def process_all(self):
        """批量处理所有文件"""
        if not self.validate_ready():
            return

        # 批量模式仍使用时间戳目录
        output_folder = self.create_output_folder(mode='batch')
        output_path = os.path.join(output_folder, "批量识别结果.xlsx")

        results = []
        for idx, img_pil in enumerate(self.images):
            result_row = self.process_single_image(img_pil)
            results.append(result_row)

        self.save_results(results, output_path)
        messagebox.showinfo("完成",
                            f"已处理{len(results)}个文件，结果保存至：\n{os.path.abspath(output_path)}")

    def preprocess_image(self, image):
        """图像预处理增强识别效果"""
        # 转为灰度图
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        # 自适应二值化
        thresh = cv2.adaptiveThreshold(gray, 255,
                                       cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY, 11, 2)
        # 降噪
        denoised = cv2.fastNlMeansDenoising(thresh, h=10)
        # 转为三通道
        return cv2.cvtColor(denoised, cv2.COLOR_GRAY2BGR)

    def validate_ready(self):
        """验证是否准备好进行识别"""
        if not self.images:
            messagebox.showwarning("警告", "请先选择需要处理的文件")
            return False
        if not self.regions:
            messagebox.showwarning("警告", "请先框选识别区域")
            return False
        return True

    def process_single_image(self, img_pil):
        """处理单个图片并返回识别结果"""
        img_cv = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)
        row = []
        for region in self.regions:
            x0, y0, x1, y1 = region
            roi = img_cv[y0:y1, x0:x1]
            roi = self.preprocess_image(roi)

            try:
                result = self.ocr.ocr(roi, cls=True)
                text = ' '.join([line[1][0] for line in result[0]]) if result[0] else ''
            except Exception as e:
                text = f"识别错误: {str(e)}"

            row.append(text)
        return row

    def run(self):
        self.root.mainloop()

    def generate_filename(self, file_info):
        """生成符合Windows规范的Excel文件名"""
        original_name = os.path.splitext(file_info['filename'])[0]

        # 1. 去除特殊字符
        clean_name = re.sub(r'[\\/*?:"<>|]', "", original_name)

        # 2. 标准化Unicode字符（处理非常规空格等）
        clean_name = unicodedata.normalize('NFKC', clean_name)

        # 3. 替换连续空格为单个下划线
        clean_name = re.sub(r'\s+', '_', clean_name)

        # 4. 添加页码信息
        if file_info['page']:
            clean_name += f"_p{file_info['page']}"

        # 5. 限制文件名长度（保留扩展名前50字符）
        max_length = 50  # 留出足够空间给时间戳和页码
        clean_name = clean_name[:max_length].rstrip('_.')

        # 6. 添加防重复后缀
        timestamp = datetime.now().strftime("%m%d%H%M%S")
        return f"{clean_name}_{timestamp}"

    def create_output_folder(self, mode='batch'):
        """创建输出目录（批量模式带时间戳，单个模式固定目录）"""
        if mode == 'single':
            # 固定目录存放所有单个结果
            output_dir = os.path.join(os.getcwd(), '单个识别结果')
            # 自动清理3天前的文件（可选）
            self.clean_old_files(output_dir, days=3)
        else:
            # 批量模式仍用时间戳目录
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.join(os.getcwd(), f'批量识别结果_{timestamp}')

        os.makedirs(output_dir, exist_ok=True)
        return output_dir

    def clean_old_files(self, dir_path, days=3):
        """自动清理旧文件（保持目录整洁）"""
        if not os.path.exists(dir_path):
            return

        now = time.time()
        for f in os.listdir(dir_path):
            file_path = os.path.join(dir_path, f)
            if os.stat(file_path).st_mtime < now - days * 86400:
                try:
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                except Exception as e:
                    print(f"清理旧文件失败：{str(e)}")

    def save_results(self, data, output_path):
        """增强的文件保存验证"""
        try:
            # 验证路径长度
            if len(output_path) > 200:
                raise ValueError("文件路径超过Windows系统限制")

            # 强制添加正确扩展名
            if not output_path.lower().endswith('.xlsx'):
                output_path += '.xlsx'

            # 创建父目录
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # 使用绝对路径
            output_path = os.path.abspath(output_path)

            # ... 原有保存逻辑保持不变 ...
            headers = ["文件名", "文件路径", "页码"] + [f"区域{i + 1}" for i in range(len(self.regions))]

            rows = []
            for idx, row in enumerate(data):
                info = self.file_info[idx] if idx < len(self.file_info) else {}
                rows.append([
                    info.get('filename', '未知文件'),
                    info.get('filepath', '未知路径'),
                    info.get('page', 'N/A'),
                    *row
                ])

            df = pd.DataFrame(rows, columns=headers)

            # 使用openpyxl引擎写入Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='识别结果')

                # 获取工作表对象进行格式设置
                worksheet = writer.sheets['识别结果']

                # 设置标题格式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

                for col in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    worksheet.column_dimensions[get_column_letter(col)].width = 25

                # 设置文本自动换行
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        except Exception as e:
            messagebox.showerror("保存失败", f"文件保存失败：{str(e)}\n尝试路径：{output_path}")

    def process_current(self):
        """处理当前单个文件"""
        if not self.validate_ready():
            return

        # 使用固定目录存放单个结果
        output_folder = self.create_output_folder(mode='single')
        file_info = self.file_info[self.current_image_index]

        # 生成带时间戳的唯一文件名
        timestamp = datetime.now().strftime("%H%M%S")
        base_name = f"{self.generate_filename(file_info)}_{timestamp}"
        output_path = os.path.join(output_folder, f"{base_name}.xlsx")

        img_pil = self.images[self.current_image_index]
        result_row = self.process_single_image(img_pil)

        self.save_results([result_row], output_path)

        # 显示完整保存路径
        messagebox.showinfo("完成",
                            f"当前文件识别结果已保存至：\n{os.path.abspath(output_path)}")

if __name__ == "__main__":
    app = InvoiceProcessorApp()
    app.run()